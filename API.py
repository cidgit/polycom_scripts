import xlrd
import xlutils 
import xlsxwriter
import HTML
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from xlutils.copy import copy
from xlrd import open_workbook
from xlwt import Workbook, easyxf
import time
import shutil
import os
import sys
import telnetlib
import select
import socket
import time
import re
import inspect
import logging


IPList = ""
import os
import sys
import lockfile  
import time
import re
fileName = __file__

def clearLog (filename):
    #print "inside clearlog"
    if os.path.exists(filename):
        #print"file exist"
        os.remove(filename)
	return
def copylog(srcfile,dstfile):
	shutil.copy2(srcfile,dstfile)
	return

def createlogger():
    function_name = inspect.stack()[1][3]
    logger = logging.getLogger(function_name)
    logger.setLevel(logging.DEBUG) #By default, logs all messages

    ch = logging.StreamHandler() #StreamHandler logs to console
    ch.setLevel(logging.ERROR)
    ch_format = logging.Formatter('%(asctime)s - %(message)s')
    ch.setFormatter(ch_format)
    logger.addHandler(ch)

    fh = logging.FileHandler("Polycom.log".format(function_name))
    fh.setLevel(logging.DEBUG)
    fh_format = logging.Formatter('%(asctime)s - %(lineno)d - %(levelname)-8s - %(message)s')
    fh.setFormatter(fh_format)
    logger.addHandler(fh)

    return logger
	
def updatelogger(f1_logger,text):
    f1_logger.info(text)
    return


def GetTestCaseForExecution ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 0)
		if ( executionstatus == "Yes" or executionstatus == "yes" ):
			tcname = r_sheet.cell_value(r_index, 1)
			list.append (tcname)
			r_index = r_index + 1
		elif ( executionstatus == "End" or executionstatus == "end" ):
			break        
		else :
			r_index = r_index + 1
	return list
GetTestCaseForExecution('C:/OpenFlow/TestCases.xls')
def GetSUTs ( tcfile , tcname, reourcefile):
    rb = open_workbook(tcfile)
    list = []
    r_sheet = rb.sheet_by_index(0)
    num_rows = r_sheet.nrows
    num_cols = r_sheet.ncols
    r_index = 1
    while r_index < num_rows:
        tc = r_sheet.cell_value(r_index, 1)
        if ( tc == tcname ):
            sutNum = r_sheet.cell_value(r_index, 2)
            r_index = num_rows        
        else :
            r_index = r_index + 1
    rb2 = open_workbook(reourcefile)
    r2_sheet = rb2.sheet_by_index(0)
    NumOfSut = int(sutNum)
    for i in range(1,NumOfSut+1):
        sutIP = r2_sheet.cell_value(i, 0)
        list.append (sutIP)
    return list

def GetTCRow (resourcefile , testname) :
    rb = open_workbook(resourcefile)
    r_sheet = rb.sheet_by_index(0)
    num_rows = r_sheet.nrows
    num_cols = r_sheet.ncols
    tcRow = ""
    r_index = 1
    while r_index < num_rows:
        tc = r_sheet.cell_value(r_index, 1)
        if ( tc == testname ):
            tcRow = r_index
            r_index = num_rows  
        else :
            r_index = r_index + 1
    return tcRow
def UpdateTcStatus (resourcefile , testname , status) :
    tcRow = GetTCRow(resourcefile , testname)
    col_index = 2
    rb = open_workbook(resourcefile)
    wb = copy(rb)
    w_sheet = wb.get_sheet(0)
    r_index = int(tcRow)
    #c_index=0
    w_sheet.write(r_index,col_index,status)
    wb.save(resourcefile)
    return 
def GetSteps( resourcefile , testname) :
    list = []
    tcRow = GetTCRow(resourcefile , testname)
    rb = open_workbook(resourcefile)
    r_sheet = rb.sheet_by_index(0)
    num_rows = r_sheet.nrows
    num_cols = r_sheet.ncols
    c_index = 3
    r_index = tcRow +1
    while r_index < num_rows:
        tcSteps = r_sheet.cell_value(r_index , c_index)
        if ( tcSteps == "End" or tcSteps == "end" ):
            r_index = num_rows
        else :
            list.append (tcSteps)
            r_index = r_index + 1
    print list
    return list
def GetStepResponse( resourcefile , testname) :
    list = []
    tcRow = GetTCRow(resourcefile , testname)
    rb = open_workbook(resourcefile)
    r_sheet = rb.sheet_by_index(0)
    num_rows = r_sheet.nrows
    num_cols = r_sheet.ncols
    c_index = 4
    r_index = tcRow +1
    while r_index < num_rows:
        tcSteps = r_sheet.cell_value(r_index , c_index)
        if ( tcSteps == "End" or tcSteps == "end" ):
            r_index = num_rows
        else :
            list.append (tcSteps)
            r_index = r_index + 1
    print list
    return list
#GetStepResponse ("testscripts.xls" , "TC1")	
def CompareConsoleOutput ( actualStr , expectedStr):
    found = actualStr.find(expectedStr)
    print found
    return found
#CompareConsoleOutput ("I am nitesh" , "notesh")
def PythonAllocate ( num , resourcefile , list ) :
	IPList1 = ""
	count = 0
	row_index = 0
	from lockfile import LockFile
	lock = LockFile(resourcefile)
	lockid = lock.is_locked()
	print lockid
	for a in xrange(1, 2):
		if lockid == False:
			lock.acquire()
			print "I have locked Resource File"
			break
		else:
			time.sleep (10)
		lockid = lock.is_locked()
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	j = r_sheet.nrows
	while row_index < j: 
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value for col_index in xrange(r_sheet.ncols)}
		length = len(list)
		if ( d['STATUS'] == "free") :
			a = " "
			w_sheet.write(row_index,col_index-2,"allocated")
			wb.save(resourcefile)
			lock.release()
			count = count + 1
			resourcename = d['IP']
			string1 = str(list)+ a + str(resourcename)
			string2 = string1.strip()
			print string2
			list = string2.split( )
			length = len(list)
			print list
			j = 1
		row_index = row_index + 1
	print "list",list
	if ( list == "" ) :
		lock.release()
	return list
def PythonfreeResource( resourcefile , list ) :
	count = 0
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	from lockfile import LockFile
	lock = LockFile(resourcefile)
	lockid = lock.is_locked()
	print lockid
	for a in xrange(1, 10):
		if lockid == False:
			lock.acquire()
			print "I have locked Resource File"
			break
		else:
			time.sleep (10)
		lockid = lock.is_locked()
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	for row_index in xrange(1, r_sheet.nrows):
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value 
			for col_index in xrange(r_sheet.ncols)} 
		if ( d['IP'] in list) :
			count = count + 1
			print "row_index", row_index
			print "col_index" , col_index-2
			w_sheet.write(row_index,col_index-2,"free")
			wb.save(resourcefile)
		if(count == len(list)+1) :
			break
	lock.release()
	return list
	
def GetUserName( resourcefile , list ) :
	count = 0
	uname = []
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	from lockfile import LockFile
	lock = LockFile(resourcefile)
	lockid = lock.is_locked()
	print lockid
	for a in xrange(1, 10):
		if lockid == False:
			lock.acquire()
			print "I have locked Resource File"
			break
		else:
			time.sleep (10)
		lockid = lock.is_locked()
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	for row_index in xrange(1, r_sheet.nrows):
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value 
			for col_index in xrange(r_sheet.ncols)} 
		if ( d['IP'] in list) :
			count = count + 1
			uname = d['Username']
			wb.save(resourcefile)
			lock.release()
		if(count == len(list)+1) :
			break
	return uname


def GetPassword( resourcefile , list ) :
	count = 0
	pwd = []
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	from lockfile import LockFile
	lock = LockFile(resourcefile)
	lockid = lock.is_locked()
	print lockid
	for a in xrange(1, 10):
		if lockid == False:
			lock.acquire()
			print "I have locked Resource File"
			break
		else:
			time.sleep (10)
		lockid = lock.is_locked()
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	for row_index in xrange(1, r_sheet.nrows):
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value 
			for col_index in xrange(r_sheet.ncols)} 
		if ( d['IP'] in list) :
			count = count + 1
			pwd = d['Password']
			wb.save(resourcefile)
			lock.release()
		if(count == len(list)+1) :
			break
	return pwd
	
def ReadTestCase( resourcefile ) :
	list = []
	row_index = 0
	from lockfile import LockFile
	lock = LockFile(resourcefile)
	lockid = lock.is_locked()
	print lockid
	for a in xrange(1, 2):
		if lockid == False:
			lock.acquire()
			print "I have locked Resource File"
			break
		else:
			time.sleep (10)
		lockid = lock.is_locked()
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	j = r_sheet.nrows
	q = r_sheet.ncols
	print col_index
	while row_index < j: 
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value for col_index in xrange(r_sheet.ncols)}
		temp = ""
		if ( d['Execution'] == "yes") :
			temp = d['TC Name']
			print temp
			list.append(temp)
			wb.save(resourcefile)
		row_index = row_index + 1
	lock.release()
	return list

def UpdateStatus( resourcefile , testname , errormsg , vernumber) :
	count = 0
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	c_index = 0
	status_col = 0
	error_col = 0
	while c_index < num_cols:
		cell_value = r_sheet.cell_value(0, c_index)
		if (cell_value == "Status"):
			status_col = c_index
		elif (cell_value == "Error" ):
			error_col = c_index
		c_index += 1
		
	for row_index in xrange(1, r_sheet.nrows):
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value 
			for col_index in xrange(r_sheet.ncols)}
		if ( d['TC Name'] == testname) :
			w_sheet.write(row_index,status_col,"FAIL")
			error_msg = "Verification "+str(vernumber)+" failed"+'\n'+"Error : "+errormsg
			w_sheet.write(row_index,error_col,error_msg)
			wb.save(resourcefile)
			row_index = num_rows
	return 

def UpdatePass( resourcefile , testname ) :
	count = 0
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	c_index = 0
	status_col = 0
	while c_index < num_cols:
		cell_value = r_sheet.cell_value(0, c_index)
		if (cell_value == "Status"):
			status_col = c_index
		c_index += 1
		
	for row_index in xrange(1, r_sheet.nrows):
		d = {keys[col_index]: r_sheet.cell(row_index, col_index).value 
			for col_index in xrange(r_sheet.ncols)} 
		if ( d['TC Name'] == testname) :
			w_sheet.write(row_index,status_col,"PASS")
			wb.save(resourcefile)
			row_index = num_rows
	return
def CreateLogFile( filename , content) :
	f = open(filename,"w")
	f.write (content)
	f.close()
	return
def writeLog( filename , content) :
	f = open(filename,"a")
	f.write (content)
	f.close()

def ExecutionSummary( resourcefile ) :
	Statuslist = ""
	P_count = 0
	F_count = 0
	T_count = 0
	NE_count = 0
	E_count = 0
	status_col = 0
	rb = open_workbook(resourcefile)
	r_sheet = rb.sheet_by_index(0)
	keys = [r_sheet.cell(0, col_index).value for col_index in xrange(r_sheet.ncols)]
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	c_index = 0
	r_index = 1
	while c_index < num_cols:
		cell_value = r_sheet.cell_value(0, c_index)
		if (cell_value == "Status"):
			status_col = c_index
		c_index += 1
	while r_index < num_rows:
		cell_value = r_sheet.cell_value(r_index,status_col)
		if (cell_value == "PASS"):
			P_count += 1
		elif (cell_value == "FAIL"):
			F_count += 1
		else:
			NE_count += 1
		r_index += 1
	E_count =P_count + F_count
	T_count = E_count + NE_count
	Statuslist = [P_count,F_count , NE_count , E_count , T_count]
	return Statuslist

def CreateSummary( filename) :
	print filename
	P_count = 0
	F_count = 0
	T_count = 0
	NE_count = 0
	E_count = 0
	rb = open_workbook(filename)
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	workbook = xlsxwriter.Workbook('ExecutionDetails.xlsx')
	worksheet2 = workbook.add_worksheet("Summary")
	worksheet = workbook.add_worksheet("Test Case")
	bold = workbook.add_format({'bold': True})
	color_format = workbook.add_format()
	color_format.set_font_color('red')
	c1_index = 0
	while c1_index < num_cols:
		cell_value = r_sheet.cell_value(0, c1_index)
		if (cell_value == "Status"):
			status_col = c1_index
		c1_index += 1
	r1_index = 1
	while r1_index < num_rows:
		cell_value = r_sheet.cell_value(r1_index,status_col)
		if (cell_value == "PASS"):
			P_count += 1
		elif (cell_value == "FAIL"):
			F_count += 1
		else:
			NE_count += 1
		r1_index += 1
	E_count =P_count + F_count
	T_count = E_count + NE_count
#-------------------------------------------------------	
	r_index = 0
	while r_index < num_rows:
		c_index = 0	
		while c_index < num_cols:
			cell_value = r_sheet.cell_value(r_index,c_index)
			if ( r_index == 0 ):
				worksheet.write(r_index,c_index,cell_value,bold)
			else :
				if (cell_value == 'FAIL'):
					worksheet.write(r_index,c_index,cell_value,color_format)
				else:
					worksheet.write(r_index,c_index,cell_value)
			c_index += 1
		r_index += 1
	
	headings = ['Status', 'Count']
	data = [
		['Pass', 'Fail', 'Not Executed', 'Total'],
		[P_count, F_count, NE_count, T_count],
	]
	worksheet2.write_row('A1', headings, bold)
	worksheet2.write_column('A2', data[0])
	worksheet2.write_column('B2', data[1])
	chart1 = workbook.add_chart({'type': 'pie'})
	chart1.add_series({
		'categories': ['Summary', 1, 0, 2, 0],
		'values':     ['Summary', 1, 1, 2, 1],
		'line':       {'color': 'red'},
	})
	chart1.set_title({'name': 'Execution Report'})
	chart1.set_style(10)
	worksheet2.insert_chart('C2', chart1, {'x_offset': 25, 'y_offset': 10})		
		
	workbook.close()
	return
def GetTestCaseForExecution1 ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 0)
		if ( executionstatus == "Yes" or executionstatus == "yes" ):
			tcname = r_sheet.cell_value(r_index, 1)
			list.append (tcname)
			r_index = r_index + 1
		elif ( executionstatus == "End" or executionstatus == "end" ):
			break        
		else :
			r_index = r_index + 1
	return list,r_index-1
	
def GetTestCasesPassed ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 3)
		if ( executionstatus == "pass" or executionstatus == "Pass"):
			tcname = r_sheet.cell_value(r_index, 1)
			list.append (tcname)
			r_index = r_index + 1        
		else :
			r_index = r_index + 1
	return list,num_rows
def GetTestCasesFailed ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 3)
		if ( executionstatus == "fail" or executionstatus == "Fail"):
			tcname = r_sheet.cell_value(r_index, 1)
			list.append (tcname)
			r_index = r_index + 1        
		else :
			r_index = r_index + 1
	return list

def GetAllTestNames ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 3)
		tcname = r_sheet.cell_value(r_index, 1)
		list.append (tcname)
		r_index = r_index + 1        
	return list
def GetAllTestStatus ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 2)
		status = r_sheet.cell_value(r_index, 3)
		if(status == "Pass" or status == "pass"):
			list.append (status)
		elif(status == "Fail" or status == "fail"):
			list.append (status)
		else:
			list.append("Not Executed")
		r_index = r_index + 1        
	return list
def GetAllTestDescription ( tcfile ) :
	rb = open_workbook(tcfile)
	list = []
	r_sheet = rb.sheet_by_index(0)
	num_rows = r_sheet.nrows
	num_cols = r_sheet.ncols
	r_index = 1
	while r_index < num_rows:
		executionstatus = r_sheet.cell_value(r_index, 2)
		status = r_sheet.cell_value(r_index, 2)
		if(status == None):
			status == "Not Executed"
		list.append (status)
		r_index = r_index + 1        
		#print "TC NAMEEEEEEEEEE",status
	return list
def CreateReport(xlsfile,logtext,htmlFileDesn):
	execRes = []
	passRes = []
	failRes = []
	TcNames = []
	TcDescr = []
	#Getting all the details
	execRes,count = GetTestCaseForExecution1(xlsfile)
	print "xlsfile",xlsfile
	#print "The total tests executed : ",len(execRes)
	Ecount = len(execRes)
	passRes,num_tests = GetTestCasesPassed(xlsfile)
	#print "passed test case names   : ",passRes
	#print "The total tests passed   : ",len(passRes)
	Pcount = len(passRes)
	#print "The total tests          : ",num_tests
	failRes = GetTestCasesFailed(xlsfile)
	#print "passed test case names   : ",failRes
	#print "The total tests failed   : ",len(failRes)
	Fcount = len(failRes)
	TcNames = GetAllTestNames(xlsfile)
	#print "All tc names : ", TcNames
	Tcount = len(TcNames)
	#print "Tcount= ",Tcount
	#print "Pcount= ",Pcount
	#print "Fcount= ",Fcount
	TnotExecuted = Tcount - Fcount - Pcount
	TcStatus = GetAllTestStatus(xlsfile)
	#print "All tc Status : ", TcStatus
	TcDescr = GetAllTestDescription(xlsfile)
	ts = time.time()
	nameForHtmlFile = htmlFileDesn + "Report" + str(ts) + ".html"
	f2 = open (nameForHtmlFile,"w")
	t = HTML.Table(header_row=['TC Passed','TC Failed', 'TC Executed','Total TCs'])
	t.rows.append([str(Pcount), str(Fcount), str(Ecount), str(Tcount)])
	htmlcode = str(t)
	f2.write("<style>\n h1 {\nborder: 2px solid DarkSlateBlue;\n}\n</style>\n")
	f2.write("<style>\n h2,p,link {\nborder: 1px solid DarkSlateBlue;\n}\n</style>\n")
	f2.write("<style>\n TR {\nbgcolor=\"#CC3399\";\n}\n</style>\n")
	f2.write("<style>\nhead {background-color:lightgray}\nh1   {color:black}\np    {color:black}\n</style>\n")
	f2.write("<style>body {\nbackground-color: Aquamarine;\n}\nh1 {\nbackground-color: yellow ;\n}\nh2 {\nbackground-color: #B0B0B0   ;\n}\nTH {\nbackground-color: #CCCC99   ;\n}\nTD {\nbackground-color: LightGreen   ;\n}\np {\nbackground-color: white\n}\n}\n</style>")
	f2.write("<h1><center>Open Flow Conformance Test Suite</center></h1>")
	f2.write("<h2><center><b>Execution Details</center></b></h2>")
	f2.write("<center>")
	f2.write(htmlcode)
	f2.write("</center>")
	
	#Creating Pie Chart for the Test Execution Report
	f2.write("<center>")
	f2.write(" <head>\n<script type=\"text/javascript\" src=\"https://www.google.com/jsapi\"></script> \
    <script type=\"text/javascript\">   \
      google.load(\"visualization\", \"1\", {packages:[\"corechart\"]}); \
      google.setOnLoadCallback(drawChart);\
      function drawChart() {\
        var data = google.visualization.arrayToDataTable([\
          ['Task', 'Test Percentage'],\
          ['Tests Passed', %(Pcount)03d],\
          ['Tests Failed', %(Fcount)03d],\
          ['Tests Not Executed', %(TnotExecuted)03d],\
        ]);\
        var options = {\
          title: 'REPORT PIE CHART :'\
        };\
        var chart = new google.visualization.PieChart(document.getElementById('piechart'));\
        chart.draw(data, options);\
      } \
    </script>\
  </head>\
  <body>\
    <div id=\"piechart\" style=\"width: 700px; height: 400px;\"></div>\
  </body>"%  {"Pcount":Pcount,"Fcount":Fcount,"TnotExecuted":TnotExecuted,})
	f2.write("</center>")
	
	f2.write("<center>")
	f2.write("<h2>Test Details</h2>")
	t2 = HTML.Table(header_row=['TC no.','TC Name','Description', 'TC Status'])
	i = 1
	while (i < num_tests):
		t2.rows.append([i,TcNames[i-1],TcDescr[i-1],TcStatus[i-1]])
		i += 1
	htmlcode = str(t2)
	f2.write(htmlcode)
	f2.write("</center>")
	
	#Reading Execution log
	textFile = open(logtext,"r")
	text = textFile.read()
	splitData = text.split("\n")
	numLines = len(splitData)
	textFile.seek(0)
	f2.write("\n<p>")
	iter = 0
	#print numLines
	while (iter < numLines):
		line = textFile.readline()
		f2.write(line)
		f2.write("<br>")
		iter += 1
	f2.write("</p>")
	f2.close()
#CreateReport(sys.argv[1],sys.argv[2],sys.argv[3])







